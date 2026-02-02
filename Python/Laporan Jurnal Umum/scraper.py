# scraper_jurnal_umum_excel.py
# Versi untuk Excel dengan menu tanggal

import sys
import json
import urllib.parse
import asyncio
import os
from datetime import datetime, timedelta, date
from playwright.async_api import async_playwright
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import config

# ===============================
# PATH
# ===============================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(BASE_DIR, f"{config.FILE_NAME}.xlsx")

# ===============================
# SAVE EXCEL
# ===============================
def save_excel(data, path):
    wb = Workbook()
    ws = wb.active
    ws.title = config.FILE_NAME[:31]

    if not data:
        wb.save(path)
        return

    def parse_excel_date(val):
        if not val:
            return ""
        for fmt in (
            "%d/%m/%Y %H.%M.%S",
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%d",
            "%Y/%m/%d",
            "%d-%m-%Y %H.%M.%S",
            "%d-%m-%Y"
        ):
            try:
                return datetime.strptime(val, fmt)
            except:
                continue
        return val

    def parse_date_for_sort(val):
        try:
            return parse_excel_date(val)
        except:
            return datetime.max

    # Sorting
    sort_key = config.SORT_BY
    sort_order = config.SORT_ORDER.lower() == "desc"

    if sort_key and sort_key in data[0]:
        def key_func(x):
            val = x.get(sort_key, "")
            if sort_key in ("tgl", "create_at") and val:
                return parse_date_for_sort(val)
            try:
                return float(val)
            except:
                return str(val)
        data_sorted = sorted(data, key=key_func, reverse=sort_order)
    else:
        data_sorted = data

    # Ambil semua keys unik
    keys = []
    for row in data_sorted:
        for k in row.keys():
            if k not in keys:
                keys.append(k)

    header = ["No."] + keys

    # Style
    header_fill = PatternFill(start_color=config.HEADER_COLOR,
                              end_color=config.HEADER_COLOR,
                              fill_type="solid")
    header_font = Font(bold=True)
    header_align = Alignment(horizontal="center", vertical="center")
    fill_even = PatternFill(start_color=config.ROW_COLOR_EVEN,
                            end_color=config.ROW_COLOR_EVEN,
                            fill_type="solid")
    fill_odd = PatternFill(start_color=config.ROW_COLOR_ODD,
                           end_color=config.ROW_COLOR_ODD,
                           fill_type="solid")

    # Build rows
    all_rows = [header]

    row_no = 0
    for row in data_sorted:
        row_no += 1
        row_values = []
        children = None

        for k in keys:
            val = row.get(k, "")

            # Format tanggal
            if k in ("tgl", "create_at") and val:
                val = parse_excel_date(val)
            # Format angka
            elif k in ("debit", "kredit") and val:
                try:
                    val = float(val)
                except:
                    val = 0

            # Expand children (nested data)
            if k == "w2ui" and isinstance(val, dict) and "children" in val:
                children = val["children"]
                val = ""
            elif isinstance(val, (dict, list)):
                val = json.dumps(val, ensure_ascii=False)

            row_values.append(val)

        all_rows.append([row_no] + row_values)

        # Process children
        if children:
            parent_faktur = row.get("faktur", "")
            parent_tgl = row.get("tgl", "")
            parent_username = row.get("username", "")
            parent_create = row.get("create_at", "")

            for child in children:
                child_values = []
                for k in keys:
                    if k == "w2ui":
                        child_values.append("")
                    else:
                        val = child.get(k, "")

                        # Fallback ke parent
                        if not val:
                            if k == "faktur":
                                val = parent_faktur
                            elif k == "tgl":
                                val = parent_tgl
                            elif k == "username":
                                val = parent_username
                            elif k == "create_at":
                                val = parent_create

                        if k in ("tgl", "create_at") and val:
                            val = parse_excel_date(val)
                        elif k in ("debit", "kredit") and val:
                            try:
                                val = float(val)
                            except:
                                val = 0
                        if isinstance(val, (dict, list)):
                            val = json.dumps(val, ensure_ascii=False)

                        child_values.append(val)

                all_rows.append([""] + child_values)

    # Write all rows
    for row_data in all_rows:
        ws.append(row_data)

    # Format header
    for col_num, col_title in enumerate(header, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    # Zebra styling + format
    for row_idx in range(2, ws.max_row + 1):
        row_num = ws.cell(row=row_idx, column=1).value
        fill = fill_even if (row_num and row_num % 2 == 0) else fill_odd

        for col_idx, col_name in enumerate(header, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = fill

            if col_name in ("debit", "kredit"):
                cell.number_format = "#,##0.00"
            elif col_name == "tgl":
                cell.number_format = "dd/mm/yyyy"
            elif col_name == "create_at":
                cell.number_format = "dd/mm/yyyy  hh.mm.ss"

    # ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Auto width
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, config.MAX_COL_WIDTH)

    wb.save(path)

# ===============================
# GET DATE RANGE
# ===============================
def get_date_range(choice):
    """Hitung date range berdasarkan pilihan menu"""
    today = datetime.now().date()

    if choice == 1:  # Hari ini
        return str(today), str(today)

    elif choice == 2:  # Kemarin
        yesterday = today - timedelta(days=1)
        return str(yesterday), str(yesterday)

    # elif choice == 3:  # 7 hari terakhir
    #     start = today - timedelta(days=7)
    #     return str(start), str(today)

    # elif choice == 4:  # 30 hari terakhir
    #     start = today - timedelta(days=30)
    #     return str(start), str(today)

    elif choice == 3:  # Bulan ini
        start = today.replace(day=1)
        return str(start), str(today)

    elif choice == 4:  # Bulan lalu
        first_day_this_month = today.replace(day=1)
        last_day_last_month = first_day_this_month - timedelta(days=1)
        first_day_last_month = last_day_last_month.replace(day=1)
        return str(first_day_last_month), str(last_day_last_month)
    
    # semester ini
    elif choice == 5:
        if today.month <= 6:
            # Semester 1
            start = today.replace(month=1, day=1)
        else:
            # Semester 2
            start = today.replace(month=7, day=1)

        return str(start), str(today)

    # semester lalu
    elif choice == 6:
        if today.month <= 6:
            # Sekarang semester 1 → semester lalu = semester 2 tahun lalu
            year = today.year - 1
            start = date(year, 7, 1)
            end = date(year, 12, 31)
        else:
            # Sekarang semester 2 → semester lalu = semester 1 tahun ini
            year = today.year
            start = date(year, 1, 1)
            end = date(year, 6, 30)

        return str(start), str(end)

    # tahun ini
    elif choice == 7:
        start = today.replace(month=1, day=1)
        return str(start), str(today)

    # tahun lalu
    elif choice == 8:
        first_day_this_year = today.replace(month=1, day=1)
        last_day_last_year = first_day_this_year - timedelta(days=1)
        first_day_last_year = last_day_last_year.replace(month=1, day=1)
        return str(first_day_last_year), str(last_day_last_year)

    # custom
    elif choice == 9:
        # akan di-handle dari argv
        return None, None


    else:
        return str(today), str(today)

# ===============================
# MAIN SCRAPER
# ===============================
async def main_async(start_date, end_date):
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context()
        page = await context.new_page()

        print("🔐 Login...")
        await page.goto(config.LOGIN_URL, timeout=30000)
        await page.fill("#username", config.USERNAME)
        await page.fill("#password", config.PASSWORD)
        await page.click("#cmdlogin")
        await page.wait_for_timeout(3000)

        print("✅ Login sukses:", page.url)

        all_records = []

        print(f"\n📅 Periode: {start_date} s/d {end_date}")

        # Buat list hari
        start = datetime.strptime(start_date, "%Y-%m-%d").date()
        end = datetime.strptime(end_date, "%Y-%m-%d").date()

        days = []
        day = start
        while day <= end:
            days.append(day)
            day += timedelta(days=1)

        print(f"📊 Total hari: {len(days)} hari\n")

        # Fetch per hari
        async def fetch_day(day):
            payload = {
                "limit": config.LIMIT,
                "offset": 0,
                **config.EXTRA_REQUEST
            }
            payload["bsearch"].update({
                "stgl_awal": day.strftime("%Y-%m-%d"),
                "stgl_akhir": day.strftime("%Y-%m-%d"),
            })
            if config.CABANG:
                payload["bsearch"]["skd_cabang"] = config.CABANG
            if config.STATUS:
                payload["bsearch"]["sstatus"] = config.STATUS

            encoded = urllib.parse.quote(json.dumps(payload))
            url = f"{config.API_BASE}?request={encoded}"
            print(f"⏳ Fetch {day}...")

            result = await page.evaluate(
                """async (u) => {
                    const r = await fetch(u, {credentials:'include'});
                    const txt = await r.text();
                    try { return JSON.parse(txt) } catch(e){ return {"raw": txt} }
                }""",
                url,
            )

            records = []
            if isinstance(result, dict):
                for k in ("records", "data", "rows", "result"):
                    if k in result and isinstance(result[k], list):
                        records = result[k]
                        break

            print(f"✅ {day}: {len(records)} data")
            return records

        # Fetch parallel (5 hari sekaligus)
        batch_size = 10
        for i in range(0, len(days), batch_size):
            batch = days[i:i+batch_size]
            tasks = [fetch_day(day) for day in batch]
            results = await asyncio.gather(*tasks)
            for records in results:
                all_records.extend(records)

        print(f"\n📊 Total records: {len(all_records)}")

        print("💾 Menyimpan ke Excel...")
        save_excel(all_records, EXCEL_PATH)

        await browser.close()
        print(f"✅ Selesai! File: {EXCEL_PATH}")

# ===============================
# MAIN
# ===============================
if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Error: Pilihan tidak valid")
        print("Usage: python scraper_jurnal_umum_excel.py <choice>")
        print("Choice: 1-8 (menu) atau 'custom start_date end_date'")
        sys.exit(1)

    try:
        choice = int(sys.argv[1])

        if choice == 9:
            if len(sys.argv) < 4:
                print("Error: Custom butuh start_date dan end_date")
                sys.exit(1)
            start_date = sys.argv[2]
            end_date = sys.argv[3]
        else:
            start_date, end_date = get_date_range(choice)


        print(f"📅 Periode: {start_date} s/d {end_date}")
        print()

        asyncio.run(main_async(start_date, end_date))

    except ValueError:
        print("Error: Pilihan harus berupa angka")
        sys.exit(1)
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
