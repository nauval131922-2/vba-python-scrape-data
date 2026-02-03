# scraper_jurnal_umum_excel.py
# Versi untuk Excel - Ambil tanggal dari cell Excel

import sys
import json
import urllib.parse
import asyncio
import os
from datetime import datetime, timedelta

from playwright.async_api import async_playwright
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

import config


# ===============================
# PATH
# ===============================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(BASE_DIR, f"{config.FILE_NAME}.xlsx")


# ===============================
# SAVE EXCEL
# ===============================
def save_excel(data, path):
    wb = Workbook()
    ws = wb.active
    ws.title = config.FILE_NAME[:31]

    if not data:
        wb.save(path)
        return

    # -------------------------------
    # Helper: parse tanggal
    # -------------------------------
    def parse_excel_date(val):
        if not val:
            return ""
        try:
            # API return: "2026-02-02T13:21:19.554344+07:00"
            return datetime.fromisoformat(val.replace("+07:00", ""))
        except:
            return val

    def parse_date_for_sort(val):
        try:
            return parse_excel_date(val)
        except Exception:
            return datetime.max

    # -------------------------------
    # Sorting
    # -------------------------------
    sort_key = config.SORT_BY
    sort_desc = config.SORT_ORDER.lower() == "desc"

    if sort_key and sort_key in data[0]:

        def key_func(x):
            val = x.get(sort_key, "")
            if sort_key == "Datetime" and val:
                return parse_date_for_sort(val)
            try:
                return float(val)
            except Exception:
                return str(val)

        data_sorted = sorted(data, key=key_func, reverse=sort_desc)
    else:
        data_sorted = data

    # -------------------------------
    # Ambil semua key unik
    # -------------------------------
    keys = []
    for row in data_sorted:
        for k in row.keys():
            if k not in keys:
                keys.append(k)

    header = ["No."] + keys

    # -------------------------------
    # Style
    # -------------------------------
    header_fill = PatternFill(
        start_color=config.HEADER_COLOR,
        end_color=config.HEADER_COLOR,
        fill_type="solid",
    )
    header_font = Font(bold=True)
    header_align = Alignment(horizontal="center", vertical="center")

    fill_even = PatternFill(
        start_color=config.ROW_COLOR_EVEN,
        end_color=config.ROW_COLOR_EVEN,
        fill_type="solid",
    )
    fill_odd = PatternFill(
        start_color=config.ROW_COLOR_ODD,
        end_color=config.ROW_COLOR_ODD,
        fill_type="solid",
    )

    # -------------------------------
    # Build rows
    # -------------------------------
    all_rows = [header]
    row_no = 0

    for row in data_sorted:
        row_no += 1
        row_values = []
        
        # ✅ LOOP untuk setiap kolom
        for k in keys:
            val = row.get(k, "")
            
            # Format tanggal untuk kolom Datetime
            if k == "Datetime" and val:
                val = parse_excel_date(val)
            # Konversi dict/list ke JSON string
            elif isinstance(val, (dict, list)):
                val = json.dumps(val, ensure_ascii=False)
            
            row_values.append(val)

        all_rows.append([row_no] + row_values)

    # -------------------------------
    # Write rows
    # -------------------------------
    for r in all_rows:
        ws.append(r)

    # Header format
    for col_num, _ in enumerate(header, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    # Zebra + format kolom
    for row_idx in range(2, ws.max_row + 1):
        row_num = ws.cell(row=row_idx, column=1).value
        fill = fill_even if (row_num and row_num % 2 == 0) else fill_odd

        for col_idx, col_name in enumerate(header, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = fill

            if col_name == "Datetime":
                cell.number_format = "dd/mm/yyyy hh:mm:ss"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Auto width
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(
            max_len + 2, config.MAX_COL_WIDTH
        )

    wb.save(path)


# ===============================
# MAIN SCRAPER
# ===============================
async def main_async(start_date, end_date):
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context()
        page = await context.new_page()

        print("🔐 Login...")
        await page.goto(config.LOGIN_URL, timeout=30000)
        await page.fill("#username", config.USERNAME)
        await page.fill("#password", config.PASSWORD)
        await page.click("#cmdlogin")
        await page.wait_for_timeout(3000)

        print("✅ Login sukses:", page.url)

        start = datetime.strptime(start_date, "%Y-%m-%d").date()
        end = datetime.strptime(end_date, "%Y-%m-%d").date()

        days = []
        d = start
        while d <= end:
            days.append(d)
            d += timedelta(days=1)

        print(f"📊 Total hari: {len(days)}")

        all_records = []

        async def fetch_day(day):
            params = {
                "bsearch[stgl_awal]": day.strftime("%d-%m-%Y"),
                "bsearch[stgl_akhir]": day.strftime("%d-%m-%Y"),
                "_": 0
            }

            query = urllib.parse.urlencode(params)
            url = f"{config.API_BASE}?{query}"

            print(f"⏳ Fetch {day}...")
            result = await page.evaluate(
                """async (u) => {
                    const r = await fetch(u, {credentials:'include'});
                    const t = await r.text();
                    try { return JSON.parse(t); }
                    catch { return {"raw": t}; }
                }""",
                url,
            )

            records = []
            if isinstance(result, dict):
                for k in ("records", "data", "rows", "result"):
                    if k in result and isinstance(result[k], list):
                        records = result[k]
                        break

            print(f"✅ {day}: {len(records)} data")
            return records

        batch_size = 10
        for i in range(0, len(days), batch_size):
            batch = days[i:i + batch_size]
            results = await asyncio.gather(*(fetch_day(d) for d in batch))
            for r in results:
                all_records.extend(r)

        print(f"📊 Total records: {len(all_records)}")
        print("💾 Menyimpan ke Excel...")
        save_excel(all_records, EXCEL_PATH)

        await browser.close()
        print(f"✅ Selesai! File: {EXCEL_PATH}")


# ===============================
# MAIN
# ===============================
if __name__ == "__main__":
    if len(sys.argv) < 4:
        print("Usage: python scraper.py 9 <start_date> <end_date>")
        print("Example: python scraper.py 9 2026-02-01 2026-02-28")
        sys.exit(1)

    try:
        # Selalu expect 3 parameter: choice(9), start_date, end_date
        start_date = sys.argv[2]
        end_date = sys.argv[3]

        asyncio.run(main_async(start_date, end_date))

    except Exception as e:
        print("Error:", e)
        import traceback
        traceback.print_exc()
        sys.exit(1)
