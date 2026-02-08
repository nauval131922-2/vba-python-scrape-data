import json
import urllib.parse
import asyncio
import os
from datetime import datetime

from playwright.async_api import async_playwright
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

import config  # ← pakai config.py kamu

# ===============================
# PATH
# ===============================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(BASE_DIR, f"{config.FILE_NAME}.xlsx")


# ===============================
# SAVE EXCEL (FINAL)
# ===============================
def save_excel(data, path):
    wb = Workbook()
    ws = wb.active
    ws.title = config.FILE_NAME[:31]

    if not data:
        wb.save(path)
        return

    # Ambil semua kolom unik (URUT STABIL)
    keys = []
    for row in data:
        for k in row.keys():
            if k not in keys:
                keys.append(k)

    header = ["No."] + keys

    # Style
    header_fill = PatternFill(start_color=config.HEADER_COLOR,
                              end_color=config.HEADER_COLOR,
                              fill_type="solid")
    header_font = Font(bold=True)
    header_align = Alignment(horizontal="center", vertical="center")

    fill_even = PatternFill(start_color=config.ROW_COLOR_EVEN,
                            end_color=config.ROW_COLOR_EVEN,
                            fill_type="solid")
    fill_odd = PatternFill(start_color=config.ROW_COLOR_ODD,
                           end_color=config.ROW_COLOR_ODD,
                           fill_type="solid")

    # Header
    ws.append(header)
    for col in range(1, len(header) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    # Data
    row_no = 0
    for row in data:
        row_no += 1
        values = []

        for k in keys:
            val = row.get(k, "")
            if isinstance(val, (dict, list)):
                val = json.dumps(val, ensure_ascii=False)
            values.append(val)

        ws.append([row_no] + values)

        fill = fill_even if row_no % 2 == 0 else fill_odd
        for col in range(1, len(header) + 1):
            ws.cell(ws.max_row, col).fill = fill

    # Freeze + Filter
    # ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Auto width
    for col in ws.columns:
        max_len = 0
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(
            max_len + 2, config.MAX_COL_WIDTH
        )

    wb.save(path)

def sort_records(records, sort_rules):
    if not sort_rules or not records:
        return records

    # Sort dari prioritas TERAKHIR ke AWAL
    # (karena Python sort itu stable)
    for field, direction in reversed(sort_rules):
        reverse = str(direction).lower() == "desc"

        records.sort(
            key=lambda x: (
                x.get(field) is None,
                x.get(field)
            ),
            reverse=reverse
        )

    return records

# ===============================
# MAIN SCRAPER
# ===============================
async def main():
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context()
        page = await context.new_page()

        print("🔐 Login...")
        await page.goto(config.LOGIN_URL, timeout=30000)
        await page.fill("#username", config.USERNAME)
        await page.fill("#password", config.PASSWORD)
        await page.click("#cmdlogin")
        await page.wait_for_timeout(3000)

        print("✅ Login sukses:", page.url)

        payload = {
            "limit": config.LIMIT,
            "offset": 0,
            **config.EXTRA_REQUEST
        }

        encoded = urllib.parse.quote(json.dumps(payload))
        url = f"{config.API_BASE}?request={encoded}"

        print("📡 Fetch API...")
        print(url)

        result = await page.evaluate(
            """async (u) => {
                const r = await fetch(u, { credentials: 'include' });
                const t = await r.text();
                try { return JSON.parse(t); }
                catch { return { raw: t }; }
            }""",
            url
        )

        records = []
        if isinstance(result, dict):
            for k in ("records", "data", "rows", "result"):
                if k in result and isinstance(result[k], list):
                    records = result[k]
                    break

        print(f"📊 Total records: {len(records)}")

        print("🔃 Sorting data (Python)...")
        records = sort_records(records, config.SORT_BY)

        print("💾 Menyimpan ke Excel...")
        save_excel(records, EXCEL_PATH)

        await browser.close()
        print(f"✅ Selesai! File: {EXCEL_PATH}")


if __name__ == "__main__":
    asyncio.run(main())
